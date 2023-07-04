import string
import sys
import json

from PyQt5 import QtCore, QtGui, QtWidgets, uic
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QApplication, QWidget, QTableWidget, QTableWidgetItem, QHeaderView, QHBoxLayout, QVBoxLayout, QDialog
from simple_colors import *
import pandas as pd
import openpyxl
import math
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers, Alignment

from azimuth import Azimuth
from fullcircle import FullCircle
from spiralcirclespiral import SpiralCircleSpiral
from spiralspiral import SpiralSpiral
from jarakbebassamping import JarakBebasSamping
# from txtconvert import TxtConvert


#------------------------Class nya Azimuth---------------------#
class NewWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Enter Coordinate Points")
        self.setWindowIcon(QtGui.QIcon('logo.ico'))

        qFormLayout = QtWidgets.QFormLayout()

        self.lineEditX = {}
        self.lineEditY = {}
        for i in range(0, int(window.txtJumlahTitikKoordinat.text())):
            #-----------X---------------
            self.lineEditX[i] = QtWidgets.QLineEdit()
            qFormLayout.addRow(QtWidgets.QLabel("X"+ str(i)), self.lineEditX[i])

            #-----------Y---------------
            self.lineEditY[i] = QtWidgets.QLineEdit()

            qFormLayout.addRow(QtWidgets.QLabel("Y" + str(i)), self.lineEditY[i])

        # self.lineEditX[0].setText("8675.12")
        # self.lineEditY[0].setText("4231.89")
        # self.lineEditX[1].setText("8100")
        # self.lineEditY[1].setText("5000")
        # self.lineEditX[2].setText("7150")
        # self.lineEditY[2].setText("5150")
        # self.lineEditX[3].setText("6489.58")
        # self.lineEditY[3].setText("5861.61")

        buttonSubmit = QtWidgets.QPushButton()
        buttonSubmit.setObjectName("submit")
        buttonSubmit.setText("Submit")
        qFormLayout.addRow(QtWidgets.QLabel(""), buttonSubmit)

        self.setLayout(qFormLayout)



        buttonSubmit.clicked.connect(self.buttonSubmitClicked)

    def buttonSubmitClicked(self):
        nCoordinate = int(window.txtJumlahTitikKoordinat.text())
        arrX = {}
        arrY = {}
        for i in range(0, nCoordinate):
            arrX[i] = float(self.lineEditX[i].text())
            arrY[i] = float(self.lineEditY[i].text())

        data = Azimuth.coor(nCoordinate, arrX, arrY)

        str_ = ''
        for i in range(0, nCoordinate - 1):
            if (i > 0):
                str_ += "PI" + str(i) + "\n"

            # str_ += "Kuadran\t\t: " + data[i]["kuadran"] + "\n"

            integer, decimal    = divmod(float(data[i]["azimuth"]), 1)
            integer2, decimal2  = divmod(float(decimal) * 60, 1)
            str_ += "Azimuth\t: " + str(int(integer)) + "\u00b0 " + str(int(integer2)) + "' " + str(round(decimal2 * 60, 1)) + '"\n'

            if (i > 0):
                # str_+= "\u0394\t\t: " + data[i]["delta"] + "\n"
                integer3, decimal3 = divmod(float(data[i]["delta"]), 1)
                integer4, decimal4 = divmod(float(decimal3) * 60, 1)
                str_ += "\u0394\t: " + str(int(integer3)) + "\u00b0 " + str(int(integer4)) + "' " + str(round(decimal4 * 60, 0)) + '"\n'

            if (i > 0):
                # print(data[i]["pi"])
                str_ += data[i]["pi"] + "\n"

            data_distance   = float(data[i]['distance'])
            # data_distance_separate = self.thousandSeparator(data_distance)
            if(i == 0):
                str_ += "\nDistance Start-PI" + str(i + 1) + " : " + str(f'{data_distance:,}') + " meter\n"
            elif (i > 0):
                str_ += "\nDistance PI" + str(i) + "-PI" + str(i + 1) + " : " + str(f'{data_distance:,}') + " meter\n"



            str_ += "\n"
        window.textEdit.setPlainText(str_)
        window.new.hide()

    def thousandSeparator(self):
        return ("{:,}".format(self))


class NewExcel(QWidget):
    def __init__(self) -> object:
        super().__init__()
        self.window_width, self.window_height = 700, 500
        self.resize(self.window_width, self.window_height)
        self.setWindowTitle("Excel File Data")
        self.setWindowIcon(QtGui.QIcon('logo.ico'))

        # msg = QMessageBox()
        # msg.setText("Hore Jalan")
        # msg.exec_()
        layout = QVBoxLayout()
        self.setLayout(layout)

        self.table = QTableWidget()
        layout.addWidget(self.table)

        self.loadExcelData()

        buttonSubmitExcel = QtWidgets.QPushButton()
        buttonSubmitExcel.setObjectName("submitExcel")
        buttonSubmitExcel.setText("Submit")
        layout.addWidget(buttonSubmitExcel)

        buttonSubmitExcel.clicked.connect(self.newnewWindow)



    def newnewWindow(self):
        self.newnew = PresentExcel()
        self.newnew.show()

    def loadExcelData(self, cell = None):
        path = window.fname[0]
        print(window.fname)

        if window.fname[0] == '':
            return

        workbook    = openpyxl.load_workbook(path)
        sheet       = workbook.active

        self.table.setRowCount(sheet.max_row)
        self.table.setColumnCount(sheet.max_column)

        self.list_values = list(sheet.values)

        # for value in list_values:
        #     print(value)
        self.table.setHorizontalHeaderLabels(self.list_values[0])


        row_index = 0
        for value_tuple in self.list_values[1:]:
            # print(value_tuple)
            col_index = 0
            for value in value_tuple:
                self.table.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1
        # self.table.setItem(1, 3, QTableWidgetItem("Hello")) #Row & Column start dari 0

    def getValue(self, rowNumber, colNumber):
        return self.list_values[rowNumber][colNumber]


class PresentExcel(QWidget):
    def __init__(self):
        global x_lurus_diputar, y_lurus_diputar, x_asli_spiral1, y_asli_spiral1, dict_tikungan, sheet, coordinate_x, coordinate_y, delta, coordinate_x_pi, coordinate_y_pi
        super().__init__()
        self.setWindowTitle("Calculation Report")
        self.setWindowIcon(QtGui.QIcon('logo.ico'))
        self.window_width, self.window_height = 700,500
        self.resize(self.window_width, self.window_height)
        self.table = QTableWidget()

        layout  = QVBoxLayout()

        path = window.fname[0]
        workbook_input = openpyxl.load_workbook(path)
        sheet = workbook_input.active



        tableWidget = QTableWidget()
        tableWidget.setColumnCount(sheet.max_row + 1)
        tableWidget.setRowCount(47) #Jangan lupa ditambah kalau menambah kolom

        #adding item in table
        tableWidget.setItem(0, 0, QTableWidgetItem('PI'))
        tableWidget.setItem(1, 0, QTableWidgetItem('Curve Type'))
        tableWidget.setItem(2, 0, QTableWidgetItem('R design'))
        tableWidget.setItem(3, 0, QTableWidgetItem('R minimum without Spiral'))
        tableWidget.setItem(4, 0, QTableWidgetItem('e design'))
        tableWidget.setItem(5, 0, QTableWidgetItem('Ls Relative Slope')) #Ls kelandaian relatif
        tableWidget.setItem(6, 0, QTableWidgetItem('Ls Lateral Offset'))
        tableWidget.setItem(7, 0, QTableWidgetItem('Ls Lateral Acceleration')) #Ls akselerasi lateral
        tableWidget.setItem(8, 0, QTableWidgetItem('Superelevation Runoff')) #Pencapaian Superelevasi
        tableWidget.setItem(9, 0, QTableWidgetItem('Ls SS'))
        tableWidget.setItem(10, 0, QTableWidgetItem('Ls min'))
        tableWidget.setItem(11, 0, QTableWidgetItem('Ls design'))
        tableWidget.setItem(12, 0, QTableWidgetItem('\u03F4s'))
        tableWidget.setItem(13, 0, QTableWidgetItem('\u03F4c'))
        tableWidget.setItem(14, 0, QTableWidgetItem('Lc'))
        tableWidget.setItem(15, 0, QTableWidgetItem('L total'))
        tableWidget.setItem(16, 0, QTableWidgetItem('Minimum Length of Horizontal Curve')) #Panjang Lengkung Horizontal Minimal
        tableWidget.setItem(17, 0, QTableWidgetItem('Xs'))
        tableWidget.setItem(18, 0, QTableWidgetItem('Ys'))
        tableWidget.setItem(19, 0, QTableWidgetItem('K'))
        tableWidget.setItem(20, 0, QTableWidgetItem('P'))
        tableWidget.setItem(21, 0, QTableWidgetItem('E'))
        tableWidget.setItem(22, 0, QTableWidgetItem('M'))
        tableWidget.setItem(23, 0, QTableWidgetItem('T'))
        tableWidget.setItem(24, 0, QTableWidgetItem('TR'))
        tableWidget.setItem(25, 0, QTableWidgetItem('T + TR'))

        tableWidget.setItem(27, 0, QTableWidgetItem('Sight Distance on Curve'))
        tableWidget.setItem(28, 0, QTableWidgetItem('PSD/SSD'))
        tableWidget.setItem(29, 0, QTableWidgetItem('Horizontal Sight Line Offset Space'))  #Jarak bebas samping tersedia
        tableWidget.setItem(30, 0, QTableWidgetItem('Centerline Inside Lane (R)')) #Radius lajur terdalam
        tableWidget.setItem(31, 0, QTableWidgetItem('Horizontal Sight Line Offset Required (M)')) #Jarak M yang dibutuhkan

        tableWidget.setItem(33, 0, QTableWidgetItem('Widening on Curve'))
        tableWidget.setItem(34, 0, QTableWidgetItem('Widening'))

        tableWidget.setItem(36, 0, QTableWidgetItem('Stationing'))
        tableWidget.setItem(37, 0, QTableWidgetItem('Start'))
        tableWidget.setItem(38, 0, QTableWidgetItem('TC/TS'))
        tableWidget.setItem(39, 0, QTableWidgetItem('SC'))
        tableWidget.setItem(40, 0, QTableWidgetItem('SS'))
        tableWidget.setItem(41, 0, QTableWidgetItem('PI'))
        tableWidget.setItem(42, 0, QTableWidgetItem('CS'))
        tableWidget.setItem(43, 0, QTableWidgetItem('CT/ST'))
        tableWidget.setItem(44, 0, QTableWidgetItem('End'))

        tableWidget.setItem(46, 0, QTableWidgetItem('Turn'))

        # ------- Kolom Satuan ------#
        tableWidget.setItem(0, 1, QTableWidgetItem('')) #PI
        tableWidget.setItem(1, 1, QTableWidgetItem('')) #Jenis Lengkung
        tableWidget.setItem(2, 1, QTableWidgetItem('m')) #R Rencana
        tableWidget.setItem(3, 1, QTableWidgetItem('m')) #R minimal tanpa Spiral
        tableWidget.setItem(4, 1, QTableWidgetItem('%')) #e rencana
        tableWidget.setItem(5, 1, QTableWidgetItem('m')) # Ls kelandaian relatif
        tableWidget.setItem(6, 1, QTableWidgetItem('m')) # Ls offset Lateral
        tableWidget.setItem(7, 1, QTableWidgetItem('m')) #ls akselerasi lateral
        tableWidget.setItem(8, 1, QTableWidgetItem('m')) #pencapaian superelevasi
        tableWidget.setItem(9, 1, QTableWidgetItem('m')) #Ls SS
        tableWidget.setItem(10, 1, QTableWidgetItem('m')) #Ls min
        tableWidget.setItem(11, 1, QTableWidgetItem('m')) #Ls rencana
        tableWidget.setItem(12, 1, QTableWidgetItem('degree')) #Tetha s
        tableWidget.setItem(13, 1, QTableWidgetItem('degree')) #Tetha c
        tableWidget.setItem(14, 1, QTableWidgetItem('m')) #Lc
        tableWidget.setItem(15, 1, QTableWidgetItem('m')) #L total
        tableWidget.setItem(16, 1, QTableWidgetItem('m')) #panjang lengkung horizontal minimal
        tableWidget.setItem(17, 1, QTableWidgetItem('m')) #Xs
        tableWidget.setItem(18, 1, QTableWidgetItem('m')) #Ys
        tableWidget.setItem(19, 1, QTableWidgetItem('m')) #K
        tableWidget.setItem(20, 1, QTableWidgetItem('m')) #P
        tableWidget.setItem(21, 1, QTableWidgetItem('m')) #E
        tableWidget.setItem(22, 1, QTableWidgetItem('m')) #M
        tableWidget.setItem(23, 1, QTableWidgetItem('m'))#T
        tableWidget.setItem(24, 1, QTableWidgetItem('m')) #TR
        tableWidget.setItem(25, 1, QTableWidgetItem('m')) #T + TR

        tableWidget.setItem(27, 1, QTableWidgetItem(''))
        tableWidget.setItem(28, 1, QTableWidgetItem('m')) # JPM/JPH
        tableWidget.setItem(29, 1, QTableWidgetItem('m')) #jarak bebas samping tersedia
        tableWidget.setItem(30, 1, QTableWidgetItem('m')) #radius lajur sebelah dalam (R)
        tableWidget.setItem(31, 1, QTableWidgetItem('m')) #jarak bebas samping dibutuhkan (M)

        tableWidget.setItem(33, 1, QTableWidgetItem(''))
        tableWidget.setItem(34, 1, QTableWidgetItem('m')) #widening

        tableWidget.setItem(36, 1, QTableWidgetItem(''))
        tableWidget.setItem(37, 1, QTableWidgetItem('')) #Start
        tableWidget.setItem(38, 1, QTableWidgetItem('')) #TC/TS
        tableWidget.setItem(39, 1, QTableWidgetItem('')) #SC
        tableWidget.setItem(40, 1, QTableWidgetItem('')) #SS
        tableWidget.setItem(41, 1, QTableWidgetItem('')) #PI
        tableWidget.setItem(42, 1, QTableWidgetItem('')) #CS
        tableWidget.setItem(43, 1, QTableWidgetItem('')) #CT/ST
        tableWidget.setItem(44, 1, QTableWidgetItem('')) #End

        tableWidget.setItem(46, 1, QTableWidgetItem(''))



        excel = NewExcel()

        dict_tikungan   = {}
        coordinate_x    = []
        coordinate_y    = []
        coordinate_x_pi = []
        coordinate_y_pi = []
        for tikungan_ke in range (1, sheet.max_row):
            dict_tikungan[tikungan_ke] = "false:"

            distance_hor            = excel.getValue(tikungan_ke, 4)
            delta_azimuth_hor       = excel.getValue(tikungan_ke, 5)
            speed_plan_hor_min      = excel.getValue(tikungan_ke, 6)
            speed_plan_hor_max      = excel.getValue(tikungan_ke, 7)
            elevation_normal_hor    = excel.getValue(tikungan_ke, 9)
            elevation_max_hor       = excel.getValue(tikungan_ke, 10)
            radius_planned_min      = excel.getValue(tikungan_ke, 11)
            radius_planned_max      = excel.getValue(tikungan_ke, 12)
            one_lane_width          = excel.getValue(tikungan_ke, 13)
            rotated_lane            = excel.getValue(tikungan_ke, 14)
            # distance_end              = excel.getValue(tikungan_ke, 15)
            ruwasja                 = excel.getValue(tikungan_ke, 17)
            rs                      = excel.getValue(tikungan_ke, 18)
            psd                     = excel.getValue(tikungan_ke, 19)
            # vehicle_type            = excel.getValue(tikungan_ke, 19)
            turn_direction          = excel.getValue(tikungan_ke, 20)

            total_lane              = excel.getValue(tikungan_ke, 8)
            option_1                = excel.getValue(tikungan_ke, 1)
            option_2                = excel.getValue(tikungan_ke, 2)
            option_3                = excel.getValue(tikungan_ke, 3)
            option                  = [option_1, option_2, option_3]

            print('\nTikungan ke ' + str(tikungan_ke))

            for opt_method in option:
                if opt_method != 'FC' and opt_method != 'SCS' and opt_method != 'SS' :
                    break

                for row in range(5, 39):  # row 5 ~ 38 RADIUS
                    if elevation_max_hor == 8:
                        worksheet_source = workbook['en2 em8']
                    elif elevation_max_hor == 6:
                        worksheet_source = workbook['en2 em6']
                    else:
                        worksheet_source = workbook['en2 em4']

                    radius_cell = worksheet_source['A' + str(row)].value
                    if radius_cell >= int(radius_planned_min) and radius_cell <= int(radius_planned_max):
                        for column in range(ord('B'), ord('M')):  # row B ~ L SPEED
                            speed_cell = worksheet_source[str(chr(column)) + '4'].value
                            if speed_cell >= int(speed_plan_hor_min) and speed_cell <= int(speed_plan_hor_max):
                                # print('Tikungan : ' + opt_method + ' --- radius : ' + str(radius_cell) + ' --- speed : ' + str(speed_cell))


                                if opt_method == 'FC':
                                    dict_tikungan[tikungan_ke] = FullCircle.fc(distance_hor, delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor, \
                                                           radius_cell, one_lane_width, rotated_lane, total_lane, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6,\
                                                           dict_7, dict_8, dict_9, dict_10, dict_11, dict_12, dict_13, dict_14, ruwasja, rs, psd, turn_direction)
                                elif opt_method == 'SCS':
                                    dict_tikungan[tikungan_ke] = SpiralCircleSpiral.scs(distance_hor, delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor, \
                                                           radius_cell, one_lane_width, rotated_lane, total_lane,\
                                                           dict_1, dict_2, dict_3, dict_4, dict_5, dict_6,
                                                           dict_7, dict_8, dict_9, dict_10, dict_11, dict_12, dict_13, dict_14, turn_direction)
                                elif opt_method == 'SS':
                                    dict_tikungan[tikungan_ke] = SpiralSpiral.ss(distance_hor, delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor, \
                                                           radius_cell, one_lane_width, rotated_lane, total_lane,\
                                                           dict_1, dict_2, dict_3, dict_4, dict_5, dict_6,
                                                           dict_7, dict_8, dict_9, dict_10, dict_11, dict_12, dict_13, dict_14, turn_direction)

                                if tikungan_ke > 1:
                                    if (dict_tikungan[tikungan_ke])[0:6] != "false:":
                                        if dict_tikungan[tikungan_ke][0] == 'FC with Elevation' and dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation':
                                            distance_curve = dict_tikungan[tikungan_ke][24] + dict_tikungan[tikungan_ke - 1][24]
                                        elif dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                                            if (dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation'):
                                                tangent_length = dict_tikungan[tikungan_ke - 1][22]
                                            else:
                                                tangent_length = dict_tikungan[tikungan_ke - 1][24]
                                            distance_curve = dict_tikungan[tikungan_ke][24] + tangent_length
                                        elif dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation':
                                            if (dict_tikungan[tikungan_ke][0] == 'FC without Elevation'):
                                                tangent_length = dict_tikungan[tikungan_ke][22]
                                            else:
                                                tangent_length = dict_tikungan[tikungan_ke][24]
                                            distance_curve = tangent_length + dict_tikungan[tikungan_ke - 1][24]
                                        elif dict_tikungan[tikungan_ke][0] == 'FC without Elevation' and dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation':
                                            distance_curve = dict_tikungan[tikungan_ke][22] + dict_tikungan[tikungan_ke - 1][22]
                                        elif dict_tikungan[tikungan_ke][0] == 'FC without Elevation':
                                            distance_curve = dict_tikungan[tikungan_ke][22] + dict_tikungan[tikungan_ke - 1][24]
                                        elif dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation':
                                            distance_curve = dict_tikungan[tikungan_ke][24] + dict_tikungan[tikungan_ke - 1][22]
                                        else:
                                            distance_curve = dict_tikungan[tikungan_ke][24] + dict_tikungan[tikungan_ke - 1][24]

                                        if distance_hor < distance_curve:
                                            dict_tikungan[tikungan_ke] = "false:Not enough space for distance"
                                            break

                                        #---------------------distance_end---------------------#
                                        if tikungan_ke - 1 == 1 :
                                            # if (dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation' and dict_tikungan[tikungan_ke][0] == 'FC with Elevation') or \
                                            #         (dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation' and dict_tikungan[tikungan_ke][0] == 'FC without Elevation'):
                                            #
                                            #     distance_end = excel.getValue(tikungan_ke - 1, 15)
                                            #     dict_tikungan[tikungan_ke - 1][27]  = distance_end #TC/TS
                                            #     dict_tikungan[tikungan_ke - 1][28]  = None #SC
                                            #     dict_tikungan[tikungan_ke - 1][29]  = None #SS
                                            #     dict_tikungan[tikungan_ke - 1][30]  = distance_end + dict_tikungan[tikungan_ke - 1][22] #PI
                                            #     dict_tikungan[tikungan_ke - 1][31]  = None #CS
                                            #     dict_tikungan[tikungan_ke - 1][32]  = distance_end + dict_tikungan[tikungan_ke - 1][13]  #CT/ST
                                            #
                                            #     dict_tikungan[tikungan_ke][27]      = dict_tikungan[tikungan_ke - 1][32] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22] #TC/TS
                                            #     dict_tikungan[tikungan_ke][28]      = None  # SC
                                            #     dict_tikungan[tikungan_ke][29]      = None  # SC
                                            #     dict_tikungan[tikungan_ke][30]      = dict_tikungan[tikungan_ke][27] + dict_tikungan[tikungan_ke][22]  # PI
                                            #     dict_tikungan[tikungan_ke][31]      = None #CS
                                            #     dict_tikungan[tikungan_ke][32]      = dict_tikungan[tikungan_ke][27] + dict_tikungan[tikungan_ke][13]  #CT/ST

                                            if dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation' or dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation' :
                                                distance_start = excel.getValue(tikungan_ke - 1, 4)  # distance Start ke PI-1
                                                distance_end = excel.getValue(tikungan_ke - 1, 16)  # distance PI terakhir ke END
                                                start = excel.getValue(tikungan_ke - 1, 15)
                                                dict_tikungan[tikungan_ke - 1][36] = start  # START
                                                dict_tikungan[tikungan_ke - 1][37] = round(start + distance_start - dict_tikungan[tikungan_ke - 1][22], 1)  # TC/TS

                                                if dict_tikungan[tikungan_ke - 1][37] < 0:
                                                    dict_tikungan[tikungan_ke] = "false:Not enough space between Start-PI1"
                                                    break

                                                dict_tikungan[tikungan_ke - 1][38] = None  # SC
                                                dict_tikungan[tikungan_ke - 1][39] = None  # SS
                                                dict_tikungan[tikungan_ke - 1][40] = round(start + distance_start, 1)  # PI
                                                dict_tikungan[tikungan_ke - 1][41] = None  # CS
                                                dict_tikungan[tikungan_ke - 1][42] = round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][13], 1)  # CT/ST

                                                if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SC
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][13], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SCS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = round(dict_tikungan[tikungan_ke][38] + dict_tikungan[tikungan_ke][13], 1)  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][41] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][39] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                            elif dict_tikungan[tikungan_ke - 1][0] == 'SCS':
                                                distance_start = excel.getValue(tikungan_ke - 1, 4)  # distance Start ke PI-1
                                                distance_end = excel.getValue(tikungan_ke - 1, 16)  # distance PI terakhir ke END
                                                start = excel.getValue(tikungan_ke - 1, 15)
                                                dict_tikungan[tikungan_ke - 1][36] = start
                                                dict_tikungan[tikungan_ke - 1][37] = round(start + distance_start - dict_tikungan[tikungan_ke - 1][22], 1)  # TC/TS

                                                if dict_tikungan[tikungan_ke - 1][37] < 0:
                                                    dict_tikungan[tikungan_ke] = "false:Not enough space between Start-PI1"
                                                    break

                                                dict_tikungan[tikungan_ke - 1][38] = round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][10], 1)  # SC
                                                dict_tikungan[tikungan_ke - 1][39] = None  # SS
                                                dict_tikungan[tikungan_ke - 1][40] = round(start + distance_start, 1)  # PI
                                                dict_tikungan[tikungan_ke - 1][41] = round(dict_tikungan[tikungan_ke - 1][38] + dict_tikungan[tikungan_ke - 1][13], 1)  # CS
                                                dict_tikungan[tikungan_ke - 1][42] = round(dict_tikungan[tikungan_ke - 1][41] + dict_tikungan[tikungan_ke - 1][10], 1)  # CT/ST

                                                if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SC
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][13], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SCS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = round(dict_tikungan[tikungan_ke][38] + dict_tikungan[tikungan_ke][13], 1)  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][41] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][39] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                            elif dict_tikungan[tikungan_ke - 1][0] == 'SS':
                                                distance_start = excel.getValue(tikungan_ke - 1, 4)  # distance Start ke PI-1
                                                distance_end = excel.getValue(tikungan_ke - 1, 16)  # distance PI terakhir ke END
                                                start = excel.getValue(tikungan_ke - 1, 15)
                                                dict_tikungan[tikungan_ke - 1][36] = start
                                                dict_tikungan[tikungan_ke - 1][37] = round(start + distance_start - dict_tikungan[tikungan_ke - 1][22], 1)  # TC/TS

                                                if dict_tikungan[tikungan_ke - 1][37] < 0:
                                                    dict_tikungan[tikungan_ke] = "false:Not enough space between Start-PI1"
                                                    break

                                                dict_tikungan[tikungan_ke - 1][38] = None  # SC
                                                dict_tikungan[tikungan_ke - 1][39] = round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][10], 1)  # SS
                                                dict_tikungan[tikungan_ke - 1][40] = round(start + distance_start, 1)  # PI
                                                dict_tikungan[tikungan_ke - 1][41] = None  # CS
                                                dict_tikungan[tikungan_ke - 1][42] = round(dict_tikungan[tikungan_ke - 1][39] + dict_tikungan[tikungan_ke - 1][10], 1)  # CT/ST

                                                if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SC
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][13], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SCS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = round(dict_tikungan[tikungan_ke][38] + dict_tikungan[tikungan_ke][13], 1)  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][41] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][39] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                        elif tikungan_ke - 1 > 1 :
                                            if dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation' or dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation' :
                                                distance_end = excel.getValue(tikungan_ke - 1, 16)
                                                dict_tikungan[tikungan_ke - 1][37] = round(dict_tikungan[tikungan_ke - 1][37], 1)  # TC/TS
                                                dict_tikungan[tikungan_ke - 1][38] = None  # SC
                                                dict_tikungan[tikungan_ke - 1][39] = None  # SS
                                                dict_tikungan[tikungan_ke - 1][40] = round(dict_tikungan[tikungan_ke - 1][40], 1) #round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][22], 1)  # PI
                                                dict_tikungan[tikungan_ke - 1][41] = None  # CS
                                                dict_tikungan[tikungan_ke - 1][42] = round(dict_tikungan[tikungan_ke - 1][42], 1) #round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][13], 1)  # CT/ST

                                                if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SC
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][13], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SCS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = round(dict_tikungan[tikungan_ke][38] + dict_tikungan[tikungan_ke][13], 1)  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][41] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][39] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                            elif dict_tikungan[tikungan_ke - 1][0] == 'SCS':
                                                distance_end = excel.getValue(tikungan_ke - 1, 16)
                                                dict_tikungan[tikungan_ke - 1][37] = round(dict_tikungan[tikungan_ke - 1][37], 1)  # TC/TS
                                                dict_tikungan[tikungan_ke - 1][38] = round(dict_tikungan[tikungan_ke - 1][38], 1) # round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][10], 1)  # SC
                                                dict_tikungan[tikungan_ke - 1][39] = None  # SS
                                                dict_tikungan[tikungan_ke - 1][40] = round(dict_tikungan[tikungan_ke - 1][40], 1) # round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][22], 1)  # PI
                                                dict_tikungan[tikungan_ke - 1][41] = round(dict_tikungan[tikungan_ke - 1][41], 1) # round(dict_tikungan[tikungan_ke - 1][38] + dict_tikungan[tikungan_ke - 1][13], 1)  # CS
                                                dict_tikungan[tikungan_ke - 1][42] = round(dict_tikungan[tikungan_ke - 1][42], 1) # round(dict_tikungan[tikungan_ke - 1][41] + dict_tikungan[tikungan_ke - 1][10], 1)  # CT/ST

                                                if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SC
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][13], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SCS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = round(dict_tikungan[tikungan_ke][38] + dict_tikungan[tikungan_ke][13], 1)  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][41] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][39] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                            elif dict_tikungan[tikungan_ke - 1][0] == 'SS':
                                                distance_end = excel.getValue(tikungan_ke - 1, 16)
                                                dict_tikungan[tikungan_ke - 1][37] = round(dict_tikungan[tikungan_ke - 1][37], 1)  # TC/TS
                                                dict_tikungan[tikungan_ke - 1][38] = None  # SC
                                                dict_tikungan[tikungan_ke - 1][39] = round(dict_tikungan[tikungan_ke - 1][39], 1) # round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][10], 1)  # SS
                                                dict_tikungan[tikungan_ke - 1][40] = round(dict_tikungan[tikungan_ke - 1][40], 1) # round(dict_tikungan[tikungan_ke - 1][37] + dict_tikungan[tikungan_ke - 1][22], 1)  # PI
                                                dict_tikungan[tikungan_ke - 1][41] = None  # CS
                                                dict_tikungan[tikungan_ke - 1][42] = round(dict_tikungan[tikungan_ke - 1][42], 1) # round(dict_tikungan[tikungan_ke - 1][39] + dict_tikungan[tikungan_ke - 1][10], 1)  # CT/ST

                                                if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SC
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][13], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SCS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SC
                                                    dict_tikungan[tikungan_ke][39] = None  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = round(dict_tikungan[tikungan_ke][38] + dict_tikungan[tikungan_ke][13], 1)  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][41] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST

                                                elif dict_tikungan[tikungan_ke][0] == 'SS' :
                                                    dict_tikungan[tikungan_ke][37] = round(dict_tikungan[tikungan_ke - 1][42] + distance_hor - dict_tikungan[tikungan_ke - 1][22] - dict_tikungan[tikungan_ke][22], 1)  # TC/TS
                                                    dict_tikungan[tikungan_ke][38] = None  # SC
                                                    dict_tikungan[tikungan_ke][39] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][10], 1)  # SS
                                                    dict_tikungan[tikungan_ke][40] = round(dict_tikungan[tikungan_ke][37] + dict_tikungan[tikungan_ke][22], 1)  # PI
                                                    dict_tikungan[tikungan_ke][41] = None  # CS
                                                    dict_tikungan[tikungan_ke][42] = round(dict_tikungan[tikungan_ke][39] + dict_tikungan[tikungan_ke][10], 1)  # CT/ST


                                        if tikungan_ke + 1 == sheet.max_row:
                                            print (dict_tikungan[tikungan_ke])
                                            distance_end = excel.getValue(tikungan_ke, 16) #distance last PI to END
                                            dict_tikungan[tikungan_ke][43] = round(dict_tikungan[tikungan_ke][40] + distance_end, 1)  # END
                                        #-----------------------------------STATIONING END---------------------------------------#


                                        #--------------------------------------- Koordinat --------------------------------------#
                                        # coordinate_x = []
                                        # coordinate_y = []
                                        # # for tikungan_ke in range(1, sheet.max_row):
                                        # #     dict_tikungan[tikungan_ke] = False
                                        # if tikungan_ke - 1 == 1:
                                        #     coordinate_x.append(0)
                                        #     coordinate_y.append(0)
                                        #     if dict_tikungan[tikungan_ke - 1][42] == 'KIRI':
                                        #         if dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation':
                                        #             fi = ((dict_tikungan[tikungan_ke - 1][13] / 5) / dict_tikungan[tikungan_ke - 1][1]) * (360 / (2 * math.pi))
                                        #             # print(dict_tikungan[tikungan_ke - 1][13])
                                        #             # print(dict_tikungan[tikungan_ke - 1][1])
                                        #             # print(math.pi)
                                        #             # print(fi)
                                        #             for i in range (1, 6):
                                        #                 x_asli = dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians(i * fi)))
                                        #                 y_asli = 2 * dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians((i * fi)/2)**2))
                                        #
                                        #                 x_diputar = x_asli
                                        #                 y_diputar = y_asli
                                        #                 coordinate_x.append(x_diputar)
                                        #                 coordinate_y.append(y_diputar)
                                        #
                                        #             print(coordinate_x)
                                        #             print(coordinate_y)
                                        #
                                        #
                                        #         elif dict_tikungan[tikungan_ke - 1][0] == 'SCS':
                                        #             a = dict_tikungan[tikungan_ke - 1][10] / 5
                                        #             fi = ((dict_tikungan[tikungan_ke - 1][13] / 5) / dict_tikungan[tikungan_ke - 1][1]) * (360 / (2 * math.pi))
                                        #             for i in range (1, 6): #spiral section
                                        #                 x_asli_spiral1 = (a * i) - (((a * i)**5) / (40 * (dict_tikungan[tikungan_ke - 1][1])**2 * (dict_tikungan[tikungan_ke - 1][10])**2))
                                        #                 y_asli_spiral1 = ((a * i)**3) / (6 * dict_tikungan[tikungan_ke - 1][1] * dict_tikungan[tikungan_ke - 1][10])
                                        #
                                        #                 x_diputar_spiral1 = x_asli_spiral1
                                        #                 y_diputar_spiral1 = y_asli_spiral1
                                        #                 coordinate_x.append(x_diputar_spiral1)
                                        #                 coordinate_y.append(y_diputar_spiral1)
                                        #
                                        #
                                        #             for i in range (1, 6): #circle section
                                        #                 x_asli = dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians(i * fi)))
                                        #                 y_asli = 2 * dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians((i * fi)/2)**2))
                                        #
                                        #                 x_diputar = x_asli * math.cos(math.radians(dict_tikungan[tikungan_ke - 1][11])) - y_asli * math.sin(math.radians(dict_tikungan[tikungan_ke - 1][11])) + coordinate_x[5]
                                        #                 y_diputar = y_asli * math.cos(math.radians(dict_tikungan[tikungan_ke - 1][11])) + x_asli * math.sin(math.radians(dict_tikungan[tikungan_ke - 1][11])) + coordinate_y[5]
                                        #                 coordinate_x.append(x_diputar)
                                        #                 coordinate_y.append(y_diputar)
                                        #
                                        #             for i in range (1, 6): #spiral section
                                        #                 x_asli_spiral2 = (a * i) - (((a * i)**5) / (40 * (dict_tikungan[tikungan_ke - 1][1])**2 * (dict_tikungan[tikungan_ke - 1][10])**2))
                                        #                 y_asli_spiral2 = ((a * i)**3) / (6 * dict_tikungan[tikungan_ke - 1][1] * dict_tikungan[tikungan_ke - 1][10])
                                        #
                                        #                 x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(delta_azimuth_hor)) + coordinate_x[10]
                                        #                 y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(delta_azimuth_hor)) + coordinate_y[10]
                                        #                 coordinate_x.append(x_diputar_spiral2)
                                        #                 coordinate_y.append(y_diputar_spiral2)
                                        #
                                        #             print(coordinate_x)
                                        #             print(coordinate_y)
                                        #
                                        #
                                        #         else:
                                        #             a = dict_tikungan[tikungan_ke - 1][10] / 5
                                        #             for i in range (1, 6):
                                        #                 x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke - 1][1]) ** 2 * (dict_tikungan[tikungan_ke - 1][10]) ** 2))
                                        #                 y_asli_spiral1 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke - 1][1] * dict_tikungan[tikungan_ke - 1][10])
                                        #
                                        #                 x_diputar_spiral1 = x_asli_spiral1
                                        #                 y_diputar_spiral1 = y_asli_spiral1
                                        #                 coordinate_x.append(x_diputar_spiral1)
                                        #                 coordinate_y.append(y_diputar_spiral1)
                                        #
                                        #             for i in range (1, 6):
                                        #                 x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke - 1][1]) ** 2 * (dict_tikungan[tikungan_ke - 1][10]) ** 2))
                                        #                 y_asli_spiral2 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke - 1][1] * dict_tikungan[tikungan_ke - 1][10])
                                        #
                                        #                 x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(0.5 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(0.5 * delta_azimuth_hor)) + coordinate_x[5]
                                        #                 y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(0.5 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(0.5 * delta_azimuth_hor)) + coordinate_y[5]
                                        #                 coordinate_x.append(x_diputar_spiral2)
                                        #                 coordinate_y.append(y_diputar_spiral2)
                                        #
                                        #             print(coordinate_x)
                                        #             print(coordinate_y)
                                        #
                                        #     elif dict_tikungan[tikungan_ke - 1][42] == 'KANAN':
                                        #         if dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation':
                                        #             fi = ((dict_tikungan[tikungan_ke - 1][13] / 5) / dict_tikungan[tikungan_ke - 1][1]) * (360 / (2 * math.pi))
                                        #             # print(dict_tikungan[tikungan_ke - 1][13])
                                        #             # print(dict_tikungan[tikungan_ke - 1][1])
                                        #             # print(math.pi)
                                        #             # print(fi)
                                        #             for i in range(1, 6):
                                        #                 x_asli = dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians(i * fi)))
                                        #                 y_asli = -1*(2 * dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians((i * fi) / 2) ** 2)))
                                        #
                                        #                 x_diputar = x_asli
                                        #                 y_diputar = y_asli
                                        #                 coordinate_x.append(x_diputar)
                                        #                 coordinate_y.append(y_diputar)
                                        #
                                        #             print(coordinate_x)
                                        #             print(coordinate_y)
                                        #
                                        #
                                        #         elif dict_tikungan[tikungan_ke - 1][0] == 'SCS':
                                        #             a = dict_tikungan[tikungan_ke - 1][10] / 5
                                        #             fi = ((dict_tikungan[tikungan_ke - 1][13] / 5) / dict_tikungan[tikungan_ke - 1][1]) * (360 / (2 * math.pi))
                                        #             for i in range(1, 6):  # spiral section
                                        #                 x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke - 1][1]) ** 2 * (dict_tikungan[tikungan_ke - 1][10]) ** 2))
                                        #                 y_asli_spiral1 = -1*(((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke - 1][1] * dict_tikungan[tikungan_ke - 1][10]))
                                        #
                                        #                 x_diputar_spiral1 = x_asli_spiral1
                                        #                 y_diputar_spiral1 = y_asli_spiral1
                                        #                 coordinate_x.append(x_diputar_spiral1)
                                        #                 coordinate_y.append(y_diputar_spiral1)
                                        #
                                        #             for i in range(1, 6):  # circle section
                                        #                 x_asli = dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians(i * fi)))
                                        #                 y_asli = -1*(2 * dict_tikungan[tikungan_ke - 1][1] * (math.sin(math.radians((i * fi) / 2) ** 2)))
                                        #
                                        #                 x_diputar = x_asli * math.cos(math.radians(-1 * dict_tikungan[tikungan_ke - 1][11])) - y_asli * math.sin(math.radians(-1 * dict_tikungan[tikungan_ke - 1][11])) + coordinate_x[5]
                                        #                 y_diputar = y_asli * math.cos(math.radians(-1 * dict_tikungan[tikungan_ke - 1][11])) + x_asli * math.sin(math.radians(-1 * dict_tikungan[tikungan_ke - 1][11])) + coordinate_y[5]
                                        #                 coordinate_x.append(x_diputar)
                                        #                 coordinate_y.append(y_diputar)
                                        #
                                        #             for i in range(1, 6):  # spiral section
                                        #                 x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke - 1][1]) ** 2 * (dict_tikungan[tikungan_ke - 1][10]) ** 2))
                                        #                 y_asli_spiral2 = -1 * (((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke - 1][1] *dict_tikungan[tikungan_ke - 1][10]))
                                        #
                                        #                 x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(-1 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(-1 * delta_azimuth_hor)) + coordinate_x[10]
                                        #                 y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(-1 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(-1 * delta_azimuth_hor)) + coordinate_y[10]
                                        #                 coordinate_x.append(x_diputar_spiral2)
                                        #                 coordinate_y.append(y_diputar_spiral2)
                                        #
                                        #             print(coordinate_x)
                                        #             print(coordinate_y)
                                        #
                                        #
                                        #         else:
                                        #             a = dict_tikungan[tikungan_ke - 1][10] / 5
                                        #             for i in range(1, 6):
                                        #                 x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke - 1][1]) ** 2 * (dict_tikungan[tikungan_ke - 1][10]) ** 2))
                                        #                 y_asli_spiral1 = -1 * ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke - 1][1] * dict_tikungan[tikungan_ke - 1][10])
                                        #
                                        #                 x_diputar_spiral1 = x_asli_spiral1
                                        #                 y_diputar_spiral1 = y_asli_spiral1
                                        #                 coordinate_x.append(x_diputar_spiral1)
                                        #                 coordinate_y.append(y_diputar_spiral1)
                                        #
                                        #             for i in range(1, 6):
                                        #                 x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke - 1][1]) ** 2 * (dict_tikungan[tikungan_ke - 1][10]) ** 2))
                                        #                 y_asli_spiral2 = -1 * ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke - 1][1] * dict_tikungan[tikungan_ke - 1][10])
                                        #
                                        #                 x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(-1 * 0.5 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(-1 * 0.5 * delta_azimuth_hor)) + coordinate_x[5]
                                        #                 y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(-1 * 0.5 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(-1 * 0.5 * delta_azimuth_hor)) + coordinate_y[5]
                                        #                 coordinate_x.append(x_diputar_spiral2)
                                        #                 coordinate_y.append(y_diputar_spiral2)
                                        #
                                        #             print(coordinate_x)
                                        #             print(coordinate_y)

                                        # else:
                                        #     delta = 0
                                        #     for i in range (1, tikungan_ke):








                            if (dict_tikungan[tikungan_ke])[0:6] != "false:":
                                print((dict_tikungan[tikungan_ke])[0:6])
                                break

                    if (dict_tikungan[tikungan_ke])[0:6] != "false:":
                        print((dict_tikungan[tikungan_ke])[0:6])
                        break

                if (dict_tikungan[tikungan_ke])[0:6] != "false:":
                    print((dict_tikungan[tikungan_ke])[0:6])
                    break

            if (dict_tikungan[tikungan_ke])[0:6] == "false:":
                print((dict_tikungan[tikungan_ke])[0:6])
                break

            #---------------------------------------- Koordinat ---------------------------------------------#
            # global coordinate_x
            # global coordinate_y
            # coordinate_x = []
            # coordinate_y = []
            # for tikungan_ke in range(1, sheet.max_row):
            #     dict_tikungan[tikungan_ke] = False

            if tikungan_ke == 1:
                sta_start       = excel.getValue(tikungan_ke, 15) #stationing start
                distance_start = excel.getValue(tikungan_ke, 4)  # distance Start ke PI-1

                coordinate_x.append(sta_start)
                coordinate_y.append(0)
                coordinate_x_pi.append(sta_start)
                coordinate_y_pi.append(0)

                panjang_lurus_start = sta_start + distance_start - dict_tikungan[tikungan_ke][22] #panjang lurus start ke TC
                panjang_lurus_start_pi = sta_start + distance_start  #panjang lurus start ke PI1 ;
                coordinate_x.append(panjang_lurus_start)
                coordinate_y.append(0)
                coordinate_x_pi.append(panjang_lurus_start_pi)
                coordinate_y_pi.append(0)

                if dict_tikungan[tikungan_ke][45] == 'LEFT':
                    if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                        fi = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))
                        # print(dict_tikungan[tikungan_ke - 1][13])
                        # print(dict_tikungan[tikungan_ke - 1][1])
                        # print(math.pi)
                        # print(fi)
                        for i in range(1, 6):
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = 2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2))

                            x_diputar = x_asli + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar = y_asli + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        print(coordinate_x)
                        print(coordinate_y)


                    elif dict_tikungan[tikungan_ke][0] == 'SCS':
                        a = dict_tikungan[tikungan_ke][10] / 5
                        fi = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))
                        for i in range(1, 6):  # spiral section
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral1 = x_asli_spiral1 + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6):  # circle section
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = 2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2))

                            x_diputar = x_asli * math.cos(math.radians(dict_tikungan[tikungan_ke][11])) - y_asli * math.sin(math.radians(dict_tikungan[tikungan_ke][11])) + coordinate_x[6]
                            y_diputar = y_asli * math.cos(math.radians(dict_tikungan[tikungan_ke][11])) + x_asli * math.sin(math.radians(dict_tikungan[tikungan_ke][11])) + coordinate_y[6]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        for i in range(1, 6):  # spiral section
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(delta_azimuth_hor)) + coordinate_x[11]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(delta_azimuth_hor)) + coordinate_y[11]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)


                    else:
                        a = dict_tikungan[tikungan_ke][10] / 5
                        for i in range(1, 6):
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral1 = x_asli_spiral1 + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6):
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(0.5 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(0.5 * delta_azimuth_hor)) + coordinate_x[6]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(0.5 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(0.5 * delta_azimuth_hor)) + coordinate_y[6]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)

                elif dict_tikungan[tikungan_ke][45] == 'RIGHT':
                    if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                        fi = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))
                        # print(dict_tikungan[tikungan_ke - 1][13])
                        # print(dict_tikungan[tikungan_ke - 1][1])
                        # print(math.pi)
                        # print(fi)
                        for i in range(1, 6):
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = -1 * (2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2)))

                            x_diputar = x_asli + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar = y_asli + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        print(coordinate_x)
                        print(coordinate_y)


                    elif dict_tikungan[tikungan_ke][0] == 'SCS':
                        a = dict_tikungan[tikungan_ke][10] / 5
                        fi = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))
                        for i in range(1, 6):  # spiral section
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = -1 * (((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10]))

                            x_diputar_spiral1 = x_asli_spiral1 + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6):  # circle section
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = -1 * (2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2)))

                            x_diputar = x_asli * math.cos(math.radians(-1 * dict_tikungan[tikungan_ke][11])) - y_asli * math.sin(math.radians(-1 * dict_tikungan[tikungan_ke][11])) + coordinate_x[6]
                            y_diputar = y_asli * math.cos(math.radians(-1 * dict_tikungan[tikungan_ke][11])) + x_asli * math.sin(math.radians(-1 * dict_tikungan[tikungan_ke][11])) + coordinate_y[6]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        for i in range(1, 6):  # spiral section
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = -1 * (((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10]))

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(-1 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(-1 * delta_azimuth_hor)) + coordinate_x[11]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(-1 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(-1 * delta_azimuth_hor)) + coordinate_y[11]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)


                    else:
                        a = dict_tikungan[tikungan_ke][10] / 5
                        for i in range(1, 6):
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = -1 * ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral1 = x_asli_spiral1 + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6):
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = -1 * ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(-1 * 0.5 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(-1 * 0.5 * delta_azimuth_hor)) + coordinate_x[6]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(-1 * 0.5 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(-1 * 0.5 * delta_azimuth_hor)) + coordinate_y[6]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)

            elif tikungan_ke > 1:
                delta = 0
                for i in range (1, tikungan_ke): #mencari nilai delta dulu
                    value = excel.getValue(i, 5)
                    if dict_tikungan[i][45] == 'LEFT':
                        delta = delta + value
                    else:
                        delta = delta - value

                    print(delta)


                panjang_lurus_pi    = abs(dict_tikungan[tikungan_ke][40] - dict_tikungan[tikungan_ke - 1][40]) #distance PI1 ke PI2 dst.

                x_lurus_diputar_pi = panjang_lurus_pi * math.cos(math.radians(delta)) - 0 + coordinate_x_pi[len(coordinate_x_pi) - 1]   #83 nilai toleransi x
                y_lurus_diputar_pi = 0 + panjang_lurus_pi * math.sin(math.radians(delta)) + coordinate_y_pi[len(coordinate_y_pi) - 1] # 0 itu nilai titik y nya, karena lurus hanya pakai x

                coordinate_x_pi.append(x_lurus_diputar_pi)
                coordinate_y_pi.append(y_lurus_diputar_pi)

                panjang_lurus = abs(dict_tikungan[tikungan_ke][37] - dict_tikungan[tikungan_ke - 1][42])
                if dict_tikungan[tikungan_ke - 1][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke - 1][0] == 'FC with Elevation':
                    x_lurus_diputar = panjang_lurus * math.cos(math.radians(delta)) - 0 + coordinate_x[len(coordinate_x)-1] # 0 itu nilai titik y nya, karena lurus hanya pakai x
                    y_lurus_diputar = 0 + panjang_lurus * math.sin(math.radians(delta)) + coordinate_y[(len(coordinate_y)-1)]

                elif dict_tikungan[tikungan_ke - 1][0] == 'SCS':
                    x_lurus_diputar = panjang_lurus * math.cos(math.radians(delta)) - 0 + coordinate_x[len(coordinate_x)-1]
                    y_lurus_diputar = 0 + panjang_lurus * math.sin(math.radians(delta)) + coordinate_y[len(coordinate_y)-1]
                    # print(coordinate_x)

                elif dict_tikungan[tikungan_ke - 1][0] == 'SS':
                    x_lurus_diputar = panjang_lurus * math.cos(math.radians(delta)) - 0 + coordinate_x[len(coordinate_x)-1]
                    y_lurus_diputar = 0 + panjang_lurus * math.sin(math.radians(delta)) + coordinate_y[len(coordinate_y)-1]

                # else:
                #     return 'false:'

                coordinate_x.append(x_lurus_diputar)
                coordinate_y.append(y_lurus_diputar)

                if dict_tikungan[tikungan_ke][45] == 'LEFT':
                    if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                        fi          = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))


                        for i in range(1, 6):
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = 2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2))

                            x_diputar = x_asli * math.cos(math.radians(delta)) - y_asli * math.sin(math.radians(delta)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar = y_asli * math.cos(math.radians(delta)) + x_asli * math.sin(math.radians(delta)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        print(coordinate_x)
                        print(coordinate_y)


                    elif dict_tikungan[tikungan_ke][0] == 'SCS':
                        a = dict_tikungan[tikungan_ke][10] / 5
                        fi = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))
                        for i in range(1, 6):  # spiral section
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral1 = x_asli_spiral1 * math.cos(math.radians(delta)) - y_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 * math.cos(math.radians(delta)) + x_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6):  # circle section
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = 2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2))

                            x_diputar = x_asli * math.cos(math.radians(delta + dict_tikungan[tikungan_ke][11])) - y_asli * math.sin(math.radians(delta + dict_tikungan[tikungan_ke][11])) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar = y_asli * math.cos(math.radians(delta + dict_tikungan[tikungan_ke][11])) + x_asli * math.sin(math.radians(delta + dict_tikungan[tikungan_ke][11])) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        for i in range(1, 6):  # spiral section
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(delta + delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(delta + delta_azimuth_hor)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(delta + delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(delta + delta_azimuth_hor)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)


                    else:
                        a = dict_tikungan[tikungan_ke][10] / 5
                        for i in range(1, 6): #spiral 1
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral1 = x_asli_spiral1 * math.cos(math.radians(delta)) - y_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 * math.cos(math.radians(delta)) + x_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6): #spiral 2
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = ((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10])

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(delta + 0.5 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(delta + 0.5 * delta_azimuth_hor)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(delta + 0.5 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(delta + 0.5 * delta_azimuth_hor)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)

                if dict_tikungan[tikungan_ke][45] == 'RIGHT':
                    if dict_tikungan[tikungan_ke][0] == 'FC without Elevation' or dict_tikungan[tikungan_ke][0] == 'FC with Elevation':
                        fi          = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))

                        for i in range(1, 6):
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = -1 * (2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2)))

                            x_diputar = x_asli * math.cos(math.radians(delta)) - y_asli * math.sin(math.radians(delta)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar = y_asli * math.cos(math.radians(delta)) + x_asli * math.sin(math.radians(delta)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        print(coordinate_x)
                        print(coordinate_y)


                    elif dict_tikungan[tikungan_ke][0] == 'SCS':
                        a = dict_tikungan[tikungan_ke][10] / 5
                        fi = ((dict_tikungan[tikungan_ke][13] / 5) / dict_tikungan[tikungan_ke][1]) * (360 / (2 * math.pi))
                        for i in range(1, 6):  # spiral section
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = -1 * (((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10]))

                            x_diputar_spiral1 = x_asli_spiral1 * math.cos(math.radians(delta)) - y_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 * math.cos(math.radians(delta)) + x_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6):  # circle section
                            x_asli = dict_tikungan[tikungan_ke][1] * (math.sin(math.radians(i * fi)))
                            y_asli = -1 * (2 * dict_tikungan[tikungan_ke][1] * (math.sin(math.radians((i * fi) / 2) ** 2)))

                            x_diputar = x_asli * math.cos(math.radians(delta + dict_tikungan[tikungan_ke][11])) - y_asli * math.sin(math.radians(delta + dict_tikungan[tikungan_ke][11])) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar = y_asli * math.cos(math.radians(delta + dict_tikungan[tikungan_ke][11])) + x_asli * math.sin(math.radians(delta + dict_tikungan[tikungan_ke][11])) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar)
                            coordinate_y.append(y_diputar)

                        for i in range(1, 6):  # spiral section
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = -1 * (((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10]))

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(delta - delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(delta - delta_azimuth_hor)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(delta - delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(delta - delta_azimuth_hor)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)


                    else:
                        a = dict_tikungan[tikungan_ke][10] / 5
                        for i in range(1, 6): #spiral 1
                            x_asli_spiral1 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral1 = -1 * (((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10]))

                            x_diputar_spiral1 = x_asli_spiral1 * math.cos(math.radians(delta)) - y_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral1 = y_asli_spiral1 * math.cos(math.radians(delta)) + x_asli_spiral1 * math.sin(math.radians(delta)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral1)
                            coordinate_y.append(y_diputar_spiral1)

                        for i in range(1, 6): #spiral 2
                            x_asli_spiral2 = (a * i) - (((a * i) ** 5) / (40 * (dict_tikungan[tikungan_ke][1]) ** 2 * (dict_tikungan[tikungan_ke][10]) ** 2))
                            y_asli_spiral2 = -1 * (((a * i) ** 3) / (6 * dict_tikungan[tikungan_ke][1] * dict_tikungan[tikungan_ke][10]))

                            x_diputar_spiral2 = x_asli_spiral2 * math.cos(math.radians(delta - 0.5 * delta_azimuth_hor)) - y_asli_spiral2 * math.sin(math.radians(delta - 0.5 * delta_azimuth_hor)) + coordinate_x[(len(coordinate_x)-1) - (i - 1)]
                            y_diputar_spiral2 = y_asli_spiral2 * math.cos(math.radians(delta - 0.5 * delta_azimuth_hor)) + x_asli_spiral2 * math.sin(math.radians(delta - 0.5 * delta_azimuth_hor)) + coordinate_y[(len(coordinate_y)-1) - (i - 1)]
                            coordinate_x.append(x_diputar_spiral2)
                            coordinate_y.append(y_diputar_spiral2)

                        print(coordinate_x)
                        print(coordinate_y)


            if tikungan_ke + 1 == sheet.max_row:
                distance_end = excel.getValue(tikungan_ke, 16)  # distance PI terakhir ke END
                delta_end = 0
                for i in range(1, tikungan_ke + 1):  # mencari nilai delta dulu
                    value = excel.getValue(i, 5)
                    if dict_tikungan[i][45] == 'LEFT':
                        delta_end = delta_end + value
                    else:
                        delta_end = delta_end - value

                print('delta end adalah : ' + str(delta_end))
                print('jarak terakhir adalah ' + str(distance_end))

                panjang_lurus_end   = distance_end - dict_tikungan[tikungan_ke][22] #panjang PI terakhir ke End
                # panjang_lurus_end_pi= dict_tikungan[tikungan_ke][40] + distance_end

                if panjang_lurus_end < 0:
                    dict_tikungan[tikungan_ke] = "false:Not enough space between the last PI-END"
                    break

                x_lurus_diputar = panjang_lurus_end * math.cos(math.radians(delta_end)) - 0 + coordinate_x[len(coordinate_x) - 1]  # 0 itu nilai titik y nya, karena lurus hanya pakai x
                y_lurus_diputar = 0 + panjang_lurus_end * math.sin(math.radians(delta_end)) + coordinate_y[(len(coordinate_y) - 1)]
                #
                # x_lurus_diputar_pi = panjang_lurus_end_pi * math.cos(math.radians(delta_end)) - 0 + coordinate_x_pi[len(coordinate_x_pi) - 1]  # 0 itu nilai titik y nya, karena lurus hanya pakai x
                # y_lurus_diputar_pi = 0 + panjang_lurus_end_pi * math.sin(math.radians(delta_end)) + coordinate_y_pi[(len(coordinate_y_pi) - 1)]

                coordinate_x.append(x_lurus_diputar)
                coordinate_y.append(y_lurus_diputar)
                coordinate_x_pi.append(x_lurus_diputar) #titik end pasti sama dengan yang lengkung
                coordinate_y_pi.append(y_lurus_diputar) #titik end pasti sama dengan yang lengkung


            print(dict_tikungan[tikungan_ke])


            # # adding calculation into table
            # for count_calc_column in range(1, sheet.max_row):  # column 1 sampai sebanyak input
            #     for count_calc_row in range(1, 43):  # row 1-42
            #         for i in range (0, 25):
            #             tableWidget.setItem(count_calc_row, count_calc_column, QTableWidgetItem(str(dict_tikungan[tikungan_ke][i])))

        if (dict_tikungan[tikungan_ke])[0:6] == 'false:':
            error_msg = QMessageBox()
            error_msg.setWindowTitle("Can't Calculate the Results !")
            error_msg.setText("PI " + str(tikungan_ke) + " is not fulfilled.\n" + str((dict_tikungan[tikungan_ke])[6:]))
            error_msg.setIcon(QMessageBox.Critical)
            error_msg.setInformativeText("Re-check your input in Ms. Excel.")
            error_msg.exec_()
        else:
            for i in range (1, sheet.max_row):
                for j in range (0, 46): #JANGAN LUPA DIUBAH KALAU MENAMBAH BARIS
                    tableWidget.setItem(j + 1, i + 1, QTableWidgetItem(str(dict_tikungan[i][j])))


        for count_pi in range (1, sheet.max_row):
            tableWidget.setItem(0, count_pi + 1, QTableWidgetItem(str(count_pi)))


        layout.addWidget(tableWidget)

        self.setLayout(layout)
        self.show()


        #----------------------------- Tombol Save -------------------------------#
        buttonSaveHasil = QtWidgets.QPushButton()
        buttonSaveHasil.setObjectName("savehasil")
        buttonSaveHasil.setText("Save .xlsx")
        layout.addWidget(buttonSaveHasil)

        buttonSaveHasil.clicked.connect(self.exportToExcel)

        buttonSaveHasil_2 = QtWidgets.QPushButton()
        buttonSaveHasil_2.setObjectName("savehasil2")
        buttonSaveHasil_2.setText("Save .txt")
        layout.addWidget(buttonSaveHasil_2)

        buttonSaveHasil_2.clicked.connect(self.exportToTxt)

    def exportToTxt(self):
        file_filter = 'Text File (*.txt)'

        rel_path = QFileDialog.getSaveFileName(parent=self, caption='Save to a Location',
                                               directory='coordinate.txt', \
                                               filter=file_filter, initialFilter='Text File (*.txt)')

        rel_path_pi = QFileDialog.getSaveFileName(parent=self, caption='Save to a Location',
                                               directory='coordinate_pi.txt', \
                                               filter=file_filter, initialFilter='Text File (*.txt)')

        file_name = 'coordinate.txt'
        file_open = open(file_name, 'w')

        file2_name  = 'coordinate_pi.txt'
        file2_open  = open(file2_name, 'w')

        if rel_path[0] == '':
            return
        else:
            for i in range(0, len(coordinate_x)):
                coordinate = str(coordinate_x[i]) + ' ' + str(coordinate_y[i]) + '\n'
                file_open.write(coordinate)

            file_open.close()

            for i in range(0, len(coordinate_x_pi)):
                coordinate = str(coordinate_x_pi[i]) + ' ' + str(coordinate_y_pi[i]) + '\n'
                file2_open.write(coordinate)

            file2_open.close()

        # with open('coordinate.txt', 'r') as fp1, \
        #     open(rel_path, 'w') as fp2:
        #     results = fp1.read()
        #     fp2.write(results)
        #
        # with open('coordinate_pi.txt', 'r') as fp3, \
        #     open(rel_path_pi, 'w') as fp4:
        #     results_pi = fp3.read()
        #     fp4.write(results_pi)


    def exportToExcel(self):
        book = Workbook()
        sheet_2 = book.active

        sheet_2.cell(row=1, column=1).value = 'PI'
        sheet_2.cell(row=2, column=1).value = 'Curve Type'
        sheet_2.cell(row=3, column=1).value = 'R design'
        sheet_2.cell(row=4, column=1).value = 'R minimum without spiral'
        sheet_2.cell(row=5, column=1).value = 'e design'
        sheet_2.cell(row=6, column=1).value = 'Ls relative slope'
        sheet_2.cell(row=7, column=1).value = 'Ls lateral offset'
        sheet_2.cell(row=8, column=1).value = 'Ls lateral acceleration'
        sheet_2.cell(row=9, column=1).value = 'Superelevation runoff'
        sheet_2.cell(row=10, column=1).value = 'Ls SS'
        sheet_2.cell(row=11, column=1).value = 'Ls minimum'
        sheet_2.cell(row=12, column=1).value = 'Ls design'
        sheet_2.cell(row=13, column=1).value = '\u03F4s'
        sheet_2.cell(row=14, column=1).value = '\u03F4c'
        sheet_2.cell(row=15, column=1).value = 'Lc'
        sheet_2.cell(row=16, column=1).value = 'L total'
        sheet_2.cell(row=17, column=1).value = 'Minimum length of horizontal curve'
        sheet_2.cell(row=18, column=1).value = 'Xs'
        sheet_2.cell(row=19, column=1).value = 'Ys'
        sheet_2.cell(row=20, column=1).value = 'K'
        sheet_2.cell(row=21, column=1).value = 'P'
        sheet_2.cell(row=22, column=1).value = 'E'
        sheet_2.cell(row=23, column=1).value = 'M'
        sheet_2.cell(row=24, column=1).value = 'T'
        sheet_2.cell(row=25, column=1).value = 'TR'
        sheet_2.cell(row=26, column=1).value = 'T + TR'

        sheet_2.cell(row=28, column=1).value = 'Sight distance on curve'
        sheet_2.cell(row=29, column=1).value = 'PSD/SSD'
        sheet_2.cell(row=30, column=1).value = 'Horizontal sight line offset space'
        sheet_2.cell(row=31, column=1).value = 'Centerline inside lane (R)'
        sheet_2.cell(row=32, column=1).value = 'Horizontal sight line offset required (M)'

        sheet_2.cell(row=34, column=1).value = 'WIdening on curve'
        sheet_2.cell(row=35, column=1).value = 'Widening'

        sheet_2.cell(row=37, column=1).value = 'Stationing'
        sheet_2.cell(row=38, column=1).value = 'Start'
        sheet_2.cell(row=39, column=1).value = 'TC/TS'
        sheet_2.cell(row=40, column=1).value = 'SC'
        sheet_2.cell(row=41, column=1).value = 'SS'
        sheet_2.cell(row=42, column=1).value = 'PI'
        sheet_2.cell(row=43, column=1).value = 'CS'
        sheet_2.cell(row=44, column=1).value = 'CT/ST'
        sheet_2.cell(row=45, column=1).value = 'End'

        sheet_2.cell(row=47, column=1).value = 'Turn'

        #--------------Unit---------------#
        sheet_2.cell(row=1, column=2).value = ''
        sheet_2.cell(row=2, column=2).value = ''
        sheet_2.cell(row=3, column=2).value = 'm'
        sheet_2.cell(row=4, column=2).value = 'm'
        sheet_2.cell(row=5, column=2).value = '%'
        sheet_2.cell(row=6, column=2).value = 'm'
        sheet_2.cell(row=7, column=2).value = 'm'
        sheet_2.cell(row=8, column=2).value = 'm'
        sheet_2.cell(row=9, column=2).value = 'm'
        sheet_2.cell(row=10, column=2).value = 'm'
        sheet_2.cell(row=11, column=2).value = 'm'
        sheet_2.cell(row=12, column=2).value = 'm'
        sheet_2.cell(row=13, column=2).value = 'degree'
        sheet_2.cell(row=14, column=2).value = 'degree'
        sheet_2.cell(row=15, column=2).value = 'm'
        sheet_2.cell(row=16, column=2).value = 'm'
        sheet_2.cell(row=17, column=2).value = 'm'
        sheet_2.cell(row=18, column=2).value = 'm'
        sheet_2.cell(row=19, column=2).value = 'm'
        sheet_2.cell(row=20, column=2).value = 'm'
        sheet_2.cell(row=21, column=2).value = 'm'
        sheet_2.cell(row=22, column=2).value = 'm'
        sheet_2.cell(row=23, column=2).value = 'm'
        sheet_2.cell(row=24, column=2).value = 'm'
        sheet_2.cell(row=25, column=2).value = 'm'
        sheet_2.cell(row=26, column=2).value = 'm'

        sheet_2.cell(row=28, column=2).value = ''
        sheet_2.cell(row=29, column=2).value = 'm'
        sheet_2.cell(row=30, column=2).value = 'm'
        sheet_2.cell(row=31, column=2).value = 'm'
        sheet_2.cell(row=32, column=2).value = 'm'

        sheet_2.cell(row=34, column=2).value = ''
        sheet_2.cell(row=35, column=2).value = 'm'

        sheet_2.cell(row=37, column=2).value = ''
        sheet_2.cell(row=38, column=2).value = ''
        sheet_2.cell(row=39, column=2).value = ''
        sheet_2.cell(row=40, column=2).value = ''
        sheet_2.cell(row=41, column=2).value = ''
        sheet_2.cell(row=42, column=2).value = ''
        sheet_2.cell(row=43, column=2).value = ''
        sheet_2.cell(row=44, column=2).value = ''
        sheet_2.cell(row=45, column=2).value = ''

        sheet_2.cell(row=47, column=2).value = ''

        for i in range(1, sheet.max_row):
            for j in range(0, 46):
                if dict_tikungan[i][j] == []:
                    sheet_2.cell(row=j + 2, column=i + 2).value = '-'

                elif dict_tikungan[i][j] == None:
                    sheet_2.cell(row=j + 2, column=i + 2).value = '-'

                else:
                    sheet_2.cell(row=j + 2, column=i + 2).value = dict_tikungan[i][j]

        fmt_style_accounting = '_(* #,##0.0_);_(* (#,##0.0);_(* "-"?_);_(@_)' #custom format supaya muncul 1 koma desimal
        for i in range(3, 36):
            for j in range(3, sheet.max_row + 3):
                sheet_2.cell(row = i, column = j).number_format = fmt_style_accounting

        fmt_style_accounting_2decimal = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"?_);_(@_)' #custom format supaya muncul 2 koma desimal untuk widening saja
        for i in range(35, 36):
            for j in range(3, sheet.max_row + 3):
                sheet_2.cell(row = i, column = j).number_format = fmt_style_accounting_2decimal

        fmt_style = '0+000'
        for i in range(38, 46):
            for j in range(3, sheet.max_row + 3):
                sheet_2.cell(row = i, column = j).number_format = fmt_style

        for count_pi in range(1, sheet.max_row):
            sheet_2.cell(row = 1, column = count_pi + 2).value = count_pi


        for i in range (1, sheet.max_row):
            for j in range (0, 46):
                sheet_2.cell(row=j + 2, column=i + 2).alignment = Alignment(horizontal = 'right')


        #Create new sheet for coordinate points
        sheet_3 = book.create_sheet('Coordinate Points')

        sheet_3.cell(row = 1, column = 1).value = 'X'
        sheet_3.cell(row = 2, column = 1).value = 'Y'
        for i in range(1, len(coordinate_x) + 1):
            sheet_3.cell(row = 1, column = i + 1).value = coordinate_x[i-1]
            sheet_3.cell(row = 2, column = i + 1).value = coordinate_y[i-1]

        sheet_3.cell(row=3, column=1).value = 'X PI'
        sheet_3.cell(row=4, column=1).value = 'Y PI'
        for i in range(1, len(coordinate_x_pi) + 1):
            sheet_3.cell(row=3, column=i + 1).value = coordinate_x_pi[i - 1]
            sheet_3.cell(row=4, column=i + 1).value = coordinate_y_pi[i - 1]


        file_filter = 'Excel File (*.xlsx *.csv *.xls)'
        response = QFileDialog.getSaveFileName(parent=self, caption = 'Save to a Location', directory = 'Calculation Report.xlsx',\
                                               filter = file_filter, initialFilter = 'Excel File (*.xlsx *.xls)')

        if response[0] == '':
            return
        else:
            book.save(response[0])

        # print(response[0])
        # book.save(response[0])



        # columnHeaders = []
        #
        # for j in range(self.table.model().columnCount()):
        #     columnHeaders.append(self.table.horizontalHeaderItem(j).text())
        #
        # df = pd.DataFrame(columns=columnHeaders)
        #
        # for row in range(self.table.rowCount()):
        #     for col in range(self.table.columnCount()):
        #         df.at[row, columnHeaders[col]] = self.table.item(row, col).text()
        #
        # df.to_excel('Hasil Convert.xlsx', index=False)




# class SaveFile:
#     def __init__(self):
#         super().__init__()

    # def exportToExcel(self):




class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('UI/jumlahKoordinat.ui', self)
        self.window_width, self.window_height = 1031, 721
        self.resize(self.window_width, self.window_height)
        self.setWindowIcon(QtGui.QIcon('logo.ico'))
        self.buttonOK.clicked.connect(self.buttonOKClicked) #OK punya nya azimuth
        self.buttonOpenFileExcel.clicked.connect(self.buttonOpenFileExcelClicked) #Tombol Open File di Horizontal

    def buttonOpenFileExcelClicked(self): #Open File Dialog Window
        #Open File Dialog
        self.fname   = QFileDialog.getOpenFileName(self, "Open Horizontal Excel File", "", "All Files (*);;Excel Worksheet (*.xlsx)")

        if self.fname: #output file name to screen
            return self.loadExcel()


        # if not self.fname:
        #     return
        # try:
        #     with open(self.fname, 'r') as f:
        #         text = f.read()
        #         self.textEdit.setPlainText(text)
        #         self.setWindowTitle(f'{self.fname}-QtNotepad')
        #
        # except Exception as e:
        #     QMessageBox.warning(self, 'Error',
        #                         f'The following error occured:\n{type(e)}: {e}')
        #     return

    def loadExcel(self):
        self.excel = NewExcel()
        self.excel.show()

        if window.fname[0] == '':
            self.excel.hide()


    def buttonOKClicked(self): #function OK nya azimuth
        # print(self.txtJumlahTitikKoordinat.text())
        if (int(self.txtJumlahTitikKoordinat.text()) < 3):
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle('Error Message')
            msg.setText('Coordinate Points Are Less Than 3')
            msg.exec_()
            # QtWidgets.QMessageBox.about(self, 'Error Message', 'Jumlah Titik Kurang dari 3') #Ini cara lain
        else:
            return self.newWindow()
            # self.getTxtJumlahTitikKoordinat()


    def newWindow(self):
        self.new = NewWindow()
        self.new.show()

if __name__ == '__main__':






#
# print ("Skripsi Program Optimasi Horizontal & Vertikal Alinyemen Jalan by Hansel D. S.")
#
#
# #------------------------------------------- PERHITUNGAN AZIMUTH --------------------------------------------------------
#
#
#
#
#
#
# #------------------------------------------- PERHITUNGAN HORIZONTAL ALINYEMEN --------------------------------------------
    workbook  = load_workbook('Tabel Rek. Geometrik Jalan.xlsx', data_only = True) #data_only buat baca formula excel dari string ke number

    dict_1  = {}
    dict_2  = {}
    dict_3  = {}
    dict_4  = {}
    dict_5  = {}
    dict_6  = {}
    dict_7  = {}
    dict_8  = {}
    dict_9  = {}
    dict_10 = {}
    dict_11 = {}
    dict_12 = {}
    dict_13 = {}
    dict_14 = {}


    for i in range(3, 14):
        table5_42 = workbook['Tabel 5.42']  # panggil worksheet
        dict_1[table5_42['A' + str(i)].value] = {4:table5_42['B' + str(i)].value, 6:table5_42['C' + str(i)].value, 8:table5_42['D' + str(i)].value}

    for i in range(4, 15):
        table5_17 = workbook['Tabel 5.17']  # panggil worksheet
        dict_2[table5_17['A' + str(i)].value] = {0:table5_17['B' + str(i)].value}

    for i in range(2, 11):
        table5_19 = workbook['Tabel 5.19']  # panggil worksheet
        dict_3[table5_19['A' + str(i)].value] = {0:table5_19['D' + str(i)].value}

    for i in range(3, 14):
        table5_18 = workbook['Tabel 5.18']  # panggil worksheet
        dict_4[table5_18['A' + str(i)].value] = {4:table5_18['B' + str(i)].value, 6:table5_18['C' + str(i)].value, 8:table5_18['D' + str(i)].value}

    for i in range(2, 8):
        table5_22 = workbook['Tabel 5.22']  # panggil worksheet
        dict_5[table5_22['A' + str(i)].value] = {0:table5_22['B' + str(i)].value}

    for i in range(2, 14):
        table5_21 = workbook['Tabel 5.21']  # panggil worksheet
        dict_6[table5_21['A' + str(i)].value] = {0:table5_21['B' + str(i)].value}

    for i in range(5, 39):
        table_en2_em8 = workbook['en2 em8']  # panggil worksheet
        dict_7[table_en2_em8['A' + str(i)].value] = {20: table_en2_em8['B' + str(i)].value, 30: table_en2_em8['C' + str(i)].value, 40: table_en2_em8['D' + str(i)].value,\
                                                    50: table_en2_em8['E' + str(i)].value, 60: table_en2_em8['F' + str(i)].value, 70: table_en2_em8['G' + str(i)].value,\
                                                    80: table_en2_em8['H' + str(i)].value, 90: table_en2_em8['I' + str(i)].value, 100: table_en2_em8['J' + str(i)].value,\
                                                    110: table_en2_em8['K' + str(i)].value, 120: table_en2_em8['L' + str(i)].value}

    for i in range(5, 39):
        table_en2_em6 = workbook['en2 em6']  # panggil worksheet
        dict_8[table_en2_em6['A' + str(i)].value] = {20: table_en2_em6['B' + str(i)].value, 30: table_en2_em6['C' + str(i)].value, 40: table_en2_em6['D' + str(i)].value,\
                                                     50: table_en2_em6['E' + str(i)].value, 60: table_en2_em6['F' + str(i)].value, 70: table_en2_em6['G' + str(i)].value,\
                                                     80: table_en2_em6['H' + str(i)].value, 90: table_en2_em6['I' + str(i)].value, 100: table_en2_em6['J' + str(i)].value,\
                                                     110: table_en2_em6['K' + str(i)].value, 120: table_en2_em6['L' + str(i)].value}

    for i in range(5, 39):
        table_en2_em4 = workbook['en2 em4']  # panggil worksheet
        dict_9[table_en2_em4['A' + str(i)].value] = {20: table_en2_em4['B' + str(i)].value, 30: table_en2_em4['C' + str(i)].value, 40: table_en2_em4['D' + str(i)].value,\
                                                    50: table_en2_em4['E' + str(i)].value, 60: table_en2_em4['F' + str(i)].value, 70: table_en2_em4['G' + str(i)].value,\
                                                    80: table_en2_em4['H' + str(i)].value, 90: table_en2_em4['I' + str(i)].value, 100: table_en2_em4['J' + str(i)].value,\
                                                    110: table_en2_em4['K' + str(i)].value, 120: table_en2_em4['L' + str(i)].value}

    for i in range(2, 13):
        table5_41 = workbook['Tabel 5.41']  # panggil worksheet
        dict_10[table5_41['A' + str(i)].value] = {0:table5_41['B' + str(i)].value}

    for i in range(4, 15):
        table5_11 = workbook['Tabel 5.11']  # panggil worksheet
        dict_11[table5_11['A' + str(i)].value] = {0:table5_11['B' + str(i)].value}

    for i in range(3, 13):
        table5_14 = workbook['Tabel 5.14']  # panggil worksheet
        dict_12[table5_14['A' + str(i)].value] = {0:table5_14['D' + str(i)].value}

    for i in range(2, 33):
        table5_43 = workbook['Tabel 5.43'] #panggil worksheet
        dict_13[table5_43['A' + str(i)].value] = {0:table5_43['B' + str(i)].value, 1:table5_43['C' + str(i)].value}

    for i in range(3, 37):
        table5_44 = workbook['Tabel 5.44']
        dict_14[table5_44['A' + str(i)].value] = {20:table5_44['B' + str(i)].value, 30:table5_44['C' + str(i)].value, 40:table5_44['D' + str(i)].value, 50:table5_44['E' + str(i)].value, \
                                                  60:table5_44['F' + str(i)].value, 70:table5_44['G' + str(i)].value, 80:table5_44['H' + str(i)].value, 90:table5_44['I' + str(i)].value, \
                                                  100:table5_44['J' + str(i)].value}


#--------------------------------------------------------- Perhitungan -----------------------------------------------------------#

    # distance_hor            = str(425)
    # delta_azimuth_hor       = str(35.43)
    # speed_plan_hor_min      = str(40)
    # speed_plan_hor_max      = str(50)
    # elevation_normal_hor    = str(-2)
    # elevation_max_hor       = str(8)
    # radius_planned_min      = str(100)
    # radius_planned_max      = str(200)
    # option_1                = 'SCS'
    # option_2                = 'FC'
    # option_3                = 'SS'
    #
    #
    #
    # if elevation_normal_hor == '-2' and elevation_max_hor == '8':
    #     for row in range (5, 39): #row 5 ~ 38
    #         # cell_radius = workbook.active.cell(row, column)
    #         worksheet_en2em8 = workbook['en2 em8']
    #         radius_cell = worksheet_en2em8['A' + str(row)].value
    #         if radius_cell >= int(radius_planned_min) and radius_cell <= int(radius_planned_max):
    #             letters = string.ascii_uppercase
    #             for column in range (ord('B'), ord('M')):  # row B ~ L
    #                 speed_cell = worksheet_en2em8[str(chr(column)) + '4'].value
    #                 if speed_cell >= int(speed_plan_hor_min) and speed_cell <= int(speed_plan_hor_max):
    #                     print(str(chr(column)) + str(row) + ' = ' + str(worksheet_en2em8[str(chr(column)) + str(row)].value))
    #                     print(speed_cell)
    #                     print(radius_cell)
    #
    #                     if option_1 == 'FC':
    #                         valid1 = FullCircle.fc(delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor,\
    #                                       radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8,dict_9, dict_10, dict_11, dict_12)
    #                         jenis_lengkung = 'FC'
    #                     elif option_1 == 'SCS':
    #                         valid1 = SpiralCircleSpiral.scs(delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor, radius_cell, dict_1,\
    #                                                dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8, dict_9, dict_10)
    #                         jenis_lengkung = 'FC'
    #                     elif option_1 == 'SS':
    #                         valid1 = SpiralSpiral.ss(delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor, \
    #                                         radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8, dict_9, dict_10)
    #                     if valid1 == False:
    #                         if option_2 != '-':
    #                             if option_2 == 'FC':
    #                                 valid2 = FullCircle.fc(delta_azimuth_hor, speed_cell, elevation_normal_hor,
    #                                               elevation_max_hor, \
    #                                               radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7,
    #                                               dict_8, dict_9, dict_10, dict_11, dict_12)
    #                             elif option_2 == 'SCS':
    #                                 valid2 = SpiralCircleSpiral.scs(delta_azimuth_hor, speed_cell, elevation_normal_hor,
    #                                                        elevation_max_hor, radius_cell, dict_1, \
    #                                                        dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8,
    #                                                        dict_9, dict_10)
    #                             elif option_2 == 'SS':
    #                                 valid2 = SpiralSpiral.ss(delta_azimuth_hor, speed_cell, elevation_normal_hor,
    #                                                 elevation_max_hor, \
    #                                                 radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7,
    #                                                 dict_8, dict_9, dict_10)
    #                         else:
    #                             print('Tidak ada opsi metode perhitungan yang cocok')
    #
    #                         if valid2 == False:
    #                             if option_3 != '-':
    #                                 if option_3 == 'FC':
    #                                     valid3 = FullCircle.fc(delta_azimuth_hor, speed_cell, elevation_normal_hor,
    #                                                   elevation_max_hor, \
    #                                                   radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7,
    #                                                   dict_8, dict_9, dict_10, dict_11, dict_12)
    #                                 elif option_3 == 'SCS':
    #                                     valid3 = SpiralCircleSpiral.scs(delta_azimuth_hor, speed_cell, elevation_normal_hor,
    #                                                            elevation_max_hor, radius_cell, dict_1, \
    #                                                            dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8,
    #                                                            dict_9, dict_10)
    #                                 elif option_3 == 'SS':
    #                                     valid3 = SpiralSpiral.ss(delta_azimuth_hor, speed_cell, elevation_normal_hor,
    #                                                     elevation_max_hor, \
    #                                                     radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7,
    #                                                     dict_8, dict_9, dict_10)
    #                             else:
    #                                 print('Tidak ada opsi metode perhitungan yang cocok')
    #
    #
    #                     # if int(speed_cell) >= int(speed_plan_hor_min) and int(speed_plan_hor_max):
    #                     #     print('Radius = ' + str(radius_cell))
    #                     #     print('Speed = ' + str(speed_cell))
    #
    #
    #
    # if elevation_normal_hor == '-2' and elevation_max_hor == '6':
    #     for row in range (5, 39): #row 5 ~ 38
    #         # cell_radius = workbook.active.cell(row, column)
    #         worksheet_en2em6 = workbook['en2 em6']
    #         radius_cell = worksheet_en2em6['A' + str(row)].value
    #         if radius_cell >= int(radius_planned_min) and radius_cell <= int(radius_planned_max):
    #             letters = string.ascii_uppercase
    #             for column in range(ord('B'), ord('M')):  # row B ~ L
    #                 speed_cell = worksheet_en2em6[str(chr(column)) + '4'].value
    #                 if speed_cell >= int(speed_plan_hor_min) and int(speed_plan_hor_max):
    #                     print(str(chr(column)) + str(row) + ' = ' + str(worksheet_en2em6[str(chr(column)) + str(row)].value))
    #                     print(speed_cell)
    #                     print(radius_cell)
    #                     FullCircle.fc(delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor,
    #                                   radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8,
    #                                   dict_9, dict_10, dict_11, dict_12)
    #
    # if elevation_normal_hor == '-2' and elevation_max_hor == '4':
    #     for row in range (5, 39): #row 5 ~ 38
    #         # cell_radius = workbook.active.cell(row, column)
    #         worksheet_en2em4 = workbook['en2 em4']
    #         radius_cell = worksheet_en2em4['A' + str(row)].value
    #         if radius_cell >= int(radius_planned_min) and radius_cell <= int(radius_planned_max):
    #             letters = string.ascii_uppercase
    #             for column in range(ord('B'), ord('M')):  # row B ~ L
    #                 speed_cell = worksheet_en2em4[str(chr(column)) + '4'].value
    #                 if speed_cell >= int(speed_plan_hor_min) and int(speed_plan_hor_max):
    #                     print(str(chr(column)) + str(row) + ' = ' + str(worksheet_en2em4[str(chr(column)) + str(row)].value))
    #                     print(speed_cell)
    #                     print(radius_cell)
    #                     FullCircle.fc(delta_azimuth_hor, speed_cell, elevation_normal_hor, elevation_max_hor,
    #                                   radius_cell, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8,
    #                                   dict_9, dict_10, dict_11, dict_12)

    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


    #
    # #--------------------------------------------------- PERHITUNGAN FULL CIRCLE ------------------------------------------------------------------------

#
# import spiralcirclespiral
# spiralcirclespiral.scs(delta_azimuth_hor, speed_plan_hor, elevation_normal_hor, elevation_max_hor, radius_planned, dict_1, dict_2, dict_3, dict_4, dict_5, dict_6, dict_7, dict_8, dict_9, dict_10)



# print(str(dict_1[int(speed_plan_hor)][int(elevation_max_hor)]) + ' Tabel 5.42')
# print(str(dict_2[int(speed_plan_hor)][0]) + ' Tabel 5.17')
# print(str(dict_3[int(speed_plan_hor)][0]) + ' Tabel 5.19')
# print(str(dict_4[int(speed_plan_hor)][int(elevation_max_hor)]) + ' Tabel 5.18')
# is_integer(rotated_lane)
# print(str(dict_6[int(speed_plan_hor)][0]) + ' Tabel 5.21')
# print(str(dict_7[int(radius_planned)][int(speed_plan_hor)]))